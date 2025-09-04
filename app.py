#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
CTA File Naming Audit — Web Service Version

- Core audit logic preserved
- GUI removed, Flask web service added
- Upload Excel file via POST to /audit endpoint
- Returns processed file with audit results

"""

from __future__ import annotations
import os, io, sys, re, tempfile, contextlib, datetime
from pathlib import Path
from typing import Optional, Tuple, List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Flask imports for web service
from flask import Flask, request, send_file, jsonify

app = Flask(__name__)

# ---------- Version banner ----------
RULESET_NAME = "CTA Rules v1.6.1 (TF override; 64-char limit; idempotent headers; CSI date-first note; Rev & leading-space notes; 9-digit date repair; ext tokens; O→0)"

# ---------- Constants & regex ----------
DATE_RE_STRICT = re.compile(r"^\d{8}$")  # YYYYMMDD

# End-anchored Rev variant (accept ., -, _, space before/within)
REV_ANY_INFIX  = re.compile(r"[_\-\.\s]*Rev[_\-\.\s]*([0-9Oo]{1,2})$", re.IGNORECASE)
CANON_REV_SUFFIX = re.compile(r"_Rev(\d{1,2})$")  # canonical at end

# Embedded Rev tokens (with/without digits) — underscore-aware
EMBEDDED_REV = re.compile(r"(?i)(?:^|[._\-\s])Rev[._\-\s]*([0-9Oo]{1,2})(?=$|[._\-\s])")
ORPHAN_EMBEDDED_REV = re.compile(r"(?i)(?:^|[._\-\s])Rev(?=$|[._\-\s])")

MULTI_UNDERS   = re.compile(r"__+")
TRAILING_UNDERS= re.compile(r"_+$")
SPACE_BEFORE_EXT = re.compile(r"\s+\.[A-Za-z0-9]+$")

# Raised to 64 per TF override package
MAX_CHARS_TOTAL = 64

IGNORE_PREFIXES = ("I-PAY", "PCO", "IDR", "NCR", "RFI", "SMTL", "UOR")

# Prefix groups
STANDARD_PREFIXES = {
    "COR","LG","MA","MM","PN","PR","REF","RP","SCH","GA","DA","DES","DV","DW","GIS","IDR",
    "PH","SOV","BL","SV","TR","VD","IGA","ORD","OC","RE","ROW","SSC","SAF","NCR","TX","UM",
    "FA","FC","SC","CL","PL","SP","SF","WR"  # SP & WR special handling below
}
REQUIRES_CSI = {"CA","CT","EN","IR","IC","MD","MT","OM","PD","SM","SD"}  # required CSI
REQUIRES_ID_ANY = {"SOP","TF","ATP","BR","CF","CS","COST","ES","LOE","FVO","PIF","PMP","PPS","PRO","LIQ","RFP","CN","NTP","SOW","PT","AU"}
REQUIRES_ID_DIGITS_DASHES = {"PE","AEA","CO","NOC","CBL","DL","FM","PO","LNTP","CI","PA"}
OPTIONAL_ID_ANY = {"QA"}  # optional any-chars

ID_LABEL = {
    "CN":"ContractNumber","NTP":"ContractNumber","SOW":"ContractNumber","PT":"Permit-ID",
    "SOP":"SOP-ID","TF":"FormID",
    "ATP":"ProjectID","BR":"ProjectID","CF":"ProjectID","CS":"ProjectID","COST":"ProjectID","ES":"ProjectID",
    "LOE":"ProjectID","FVO":"ProjectID","PIF":"ProjectID","PMP":"ProjectID","PPS":"ProjectID",
    "PRO":"ProjectID","LIQ":"ProjectID","RFP":"ProjectID",
    "PE":"EvaluationNumber","AEA":"AllowanceNumber","CO":"ChangeOrder-Number","NOC":"ClaimNumber",
    "CBL":"BulletinNumber","DL":"BulletinNumber","FM":"FieldMemo-ID","PO":"ProceedOrder-Number",
    "LNTP":"OrderNumber","CI":"InvoiceNumber","PA":"InvoiceNumber",
    "AU":"Audit-ID","QA":"QMS-ID"
}

FILE_HEADER_LABELS = ("File Name","Filename","Document Name","file_name")
COMPANY_LABELS     = ("Company","Owner","Organization")

# Dot styling
GREEN_DOT_FONT = Font(bold=True, color="FF00B050")
RED_DOT_FONT   = Font(bold=True, color="FFFF0000")
DOT_CHAR = "●"  # colored by font

# Known extensions to strip if they sneak into title
EXT_TOKENS = {
    "pdf","doc","docx","xls","xlsx","ppt","pptx",
    "jpg","jpeg","png","tif","tiff","gif","bmp","heic",
    "dwg","rvt","csv","txt","rtf","zip","msg","eml"
}

# ---------- Helpers ----------
def safe_str(x) -> str:
    if isinstance(x, str): return x.strip()
    try:
        if pd.isna(x): return ""
    except Exception:
        pass
    return "" if x is None else str(x).strip()

def strip_spurious_extensions(text: str, notes: List[str]) -> str:
    s = text
    s2 = re.sub(
        r'(?i)(?:^|[._-])(pdf|docx?|xlsx?|pptx?|jpe?g|png|tiff?|tif|gif|bmp|heic|dwg|rvt|csv|txt|rtf|zip|msg|eml)(?=$|[._-])',
        '',
        s
    )
    lowered = s2.lower()
    for ext in EXT_TOKENS:
        if lowered.endswith(ext):
            s2 = s2[: -len(ext)]
            break
    if s2 != s:
        notes.append("Removed extra extension token from title.")
    s2 = re.sub(r'[._-]{2,}', '-', s2)
    s2 = re.sub(r'^[._-]+|[._-]+$', '', s2)
    return s2

def try_normalize_date(tok: str) -> Tuple[Optional[str], Optional[str]]:
    s = tok.strip()
    if not s:
        return None, None
    if DATE_RE_STRICT.match(s):
        return s, None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}{mo}{d}", "Normalized date from YYYY-MM-DD"
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        mo, d, y = m.groups()
        return f"{y}{int(mo):02d}{int(d):02d}", "Normalized date from M/D/YYYY"
    m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}{int(mo):02d}{int(d):02d}", "Normalized date from YYYY/M/D"
    # 9-digit malformed YYYY0MMDD → drop the stray '0' when it yields a valid date
    m = re.match(r"^(\d{4})0(\d{2})(\d{2})$", s)
    if m:
        y, mo, d = m.groups()
        try:
            datetime.date(int(y), int(mo), int(d))
            return f"{y}{mo}{d}", "Normalized date from malformed token (dropped stray 0)."
        except Exception:
            pass
    return None, None

def normalize_csi(token: str) -> Tuple[Optional[str], Optional[str]]:
    t = token.strip()
    if not t:
        return None, None
    if re.match(r"^\d{2}-\d{2}-\d{2}$", t) or re.match(r"^\d{2}-\d{2}-\d{2}-\d{2}$", t):
        return t, None
    digits = re.sub(r"\D", "", t)
    if len(digits) == 6:
        return f"{digits[0:2]}-{digits[2:4]}-{digits[4:6]}", "Auto-formatted CSI from 6 digits"
    if len(digits) == 8:
        return f"{digits[0:2]}-{digits[2:4]}-{digits[4:6]}-{digits[6:8]}", "Auto-formatted CSI from 8 digits"
    if "-" in t and len(digits) in (6,8):
        if len(digits) == 6:
            return f"{digits[0:2]}-{digits[2:4]}-{digits[4:6]}", "Corrected malformed CSI hyphenation"
        else:
            return f"{digits[0:2]}-{digits[2:4]}-{digits[4:6]}-{digits[6:8]}", "Corrected malformed CSI hyphenation"
    return None, None

def sanitize_title(raw: str, notes: List[str]) -> str:
    before = raw
    s = strip_spurious_extensions(raw, notes)  # early
    if "&" in s:
        s = s.replace("&", "and"); notes.append("Replaced '&' with 'and' in title.")
    if " " in s:
        s = re.sub(r"\s+", "", s); notes.append("Removed spaces from title.")
    if "_" in s:
        s = s.replace("_", ""); notes.append("Removed internal underscores from title.")
    s2 = re.sub(r"[^A-Za-z0-9\-]", "", s)
    if s2 != s:
        notes.append("Removed special characters from title.")
    s = s2
    # Keep CWO-# as-is (no forced change to CWO#)
    s = strip_spurious_extensions(s, notes)  # second pass
    if s != before and "Normalized title." not in notes:
        notes.append("Normalized title.")
    return s

def sanitize_id_any(raw: str, notes: List[str], label="ID") -> str:
    s = raw
    if " " in s:
        s = re.sub(r"\s+", "", s); notes.append(f"Removed spaces from {label}.")
    if "_" in s:
        s = s.replace("_", ""); notes.append(f"Removed underscores from {label}.")
    s2 = re.sub(r"[^A-Za-z0-9\-]", "", s)
    if s2 != s:
        notes.append(f"Removed special characters from {label}.")
    return s2

def sanitize_id_digits_dashes(raw: str, notes: List[str], label="Number") -> Optional[str]:
    cleaned = re.sub(r"\s+", "", raw)
    if "_" in cleaned:
        cleaned = cleaned.replace("_", ""); notes.append(f"Removed underscores from {label}.")
    digits_dashes = re.sub(r"[^0-9\-]", "", cleaned)
    if digits_dashes and re.match(r"^[0-9][0-9\-]*$", digits_dashes):
        if digits_dashes != raw:
            notes.append(f"Corrected {label} to digits/dashes only.")
        return digits_dashes
    return None

def extract_rev(trailing: str, notes: List[str]) -> Tuple[str, str]:
    """
    Normalize to exactly ONE final 'Rev##' (capital R).
    - If canonical "_Rev##" at end: keep digits, strip from trailing
    - Else if a Rev variant is at end: normalize digits (O→0), strip from trailing
    - Else strip any embedded Rev tokens (remember last digits); strip orphan 'Rev' too
      (do NOT carry orphan 'Rev' into title)
    - If no digits seen: Rev0
    Also note an extra '.' after Rev if present.

    NOTE: We intentionally preserve LEADING whitespace for sanitize_title() to record
    'Removed spaces from title.' (only trimming the RIGHT side here).
    """
    # Preserve leading spaces; only trim right side
    s = trailing.rstrip()

    had_trailing_dot_after_rev = bool(re.search(r"(?i)Rev[._\-\s]*[0-9Oo]{1,2}\.$", s))

    # Canonical suffix already present
    m_canon = CANON_REV_SUFFIX.search(s)
    if m_canon:
        digits = m_canon.group(1)[:2]
        base = s[:m_canon.start()].rstrip("_- .")
        return base, f"Rev{digits}"

    # Any Rev variant at the end
    m_end = REV_ANY_INFIX.search(s)
    captured_digits: Optional[str] = None
    if m_end:
        raw = m_end.group(1)
        digits = raw.upper().replace('O', '0')[:2]
        if digits != raw:
            notes.append("Corrected letter 'O' to digit '0' in revision.")
        captured_digits = digits
        s = s[:m_end.start()].rstrip("_- .")
        notes.append("Normalized revision element.")
    else:
        # Remove embedded Rev tokens (with digits)
        saw_embedded = False
        def _strip_embedded(m: re.Match) -> str:
            nonlocal captured_digits, saw_embedded
            saw_embedded = True
            raw = m.group(1) if m.lastindex else None
            if raw:
                d = raw.upper().replace('O', '0')[:2]
                captured_digits = d
            return ""
        s2 = EMBEDDED_REV.sub(_strip_embedded, s)
        if s2 == s:
            # Orphan 'Rev' (no digits) — drop it, don't carry into title
            s2 = ORPHAN_EMBEDDED_REV.sub("", s)
            if s2 != s:
                saw_embedded = True
                # orphan will be normalized to Rev0 below
        # Only trim trailing separators; keep leading spaces for title notes
        s = s2.rstrip("_- .")
        if saw_embedded:
            notes.append("Removed embedded revision token from title.")
            notes.append("Normalized revision element.")

    if not captured_digits:
        captured_digits = "0"
        notes.append("Added Rev0 (missing revision element).")

    if had_trailing_dot_after_rev and "Removed extra period after revision element." not in notes:
        notes.append("Removed extra period after revision element.")

    return s, f"Rev{captured_digits}"

def enforce_stem_limit(prefix: str, id_part: Optional[str], date: str, title: str, rev: str, notes: List[str]) -> str:
    base = f"{prefix}_"
    if id_part:
        base += f"{id_part}_"
    base += f"{date}_"
    stem_len = len(base) + len(title) + 1 + len(rev)  # +1 underscore before Rev
    if stem_len > MAX_CHARS_TOTAL:
        over = stem_len - MAX_CHARS_TOTAL
        keep = max(0, len(title) - over)
        title = title[:keep]
        notes.append(f"File name exceeds {MAX_CHARS_TOTAL}-character limit (title truncated).")
    return title

def join_tokens(prefix: str, id_part: Optional[str], date: str, title: str, rev: str) -> str:
    parts = [prefix]
    if id_part:
        parts.append(id_part)
    parts.extend([date, title, rev])
    s = "_".join(parts)
    s = MULTI_UNDERS.sub("_", s)
    s = TRAILING_UNDERS.sub("", s)
    return s

def underscores_ok(stem: str, has_id: bool) -> bool:
    return stem.count("_") == (4 if has_id else 3)

def postprocess_before_ext(stem: str, notes: List[str]) -> str:
    if stem.endswith("_"):
        stem = stem.rstrip("_")
        if "Removed trailing underscore before extension." not in notes:
            notes.append("Removed trailing underscore before extension.")
    if stem.endswith("."):
        stem = stem.rstrip(".")
        if "Removed extra period after revision element." not in notes:
            notes.append("Removed extra period after revision element.")
    return stem

# ---------- Layout detection ----------
def find_last_row(ws, col_idx: int, header_row: int) -> int:
    col_letter = get_column_letter(col_idx)
    r = header_row + 1
    last = header_row
    empty_streak = 0
    while r <= ws.max_row + 200:
        v = ws[f"{col_letter}{r}"].value
        if v is None or str(v).strip() == "":
            empty_streak += 1
            if empty_streak >= 5: break
        else:
            empty_streak = 0
            last = r
        r += 1
    return last

def detect_layout(ws,
                  file_labels=FILE_HEADER_LABELS,
                  company_labels=COMPANY_LABELS,
                  search_rows=60, search_cols=60):
    found = None
    labelset = {s.strip().lower() for s in file_labels}
    for r in range(1, min(search_rows, ws.max_row) + 1):
        for c in range(1, min(search_cols, ws.max_column) + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().lower() in labelset:
                found = (r, c); break
        if found: break
    if found:
        header_row, file_col = found
    else:
        header_row, file_col = 6, 2

    company_col = None
    comp_labels = {s.strip().lower() for s in company_labels}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip().lower() in comp_labels:
            company_col = c; break

    # Track last header column with any value
    last_header_col = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None or str(v).strip() == "": continue
        last_header_col = c

    # Idempotent header detection/reuse
    existing_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str):
            existing_map[v.strip().lower()] = c

    changed_col = existing_map.get("file name ok")
    status_col  = existing_map.get("status")
    newname_col = existing_map.get("new file name")
    notes_col   = existing_map.get("notes")

    # Allocate any missing audit columns to the right of last header col
    def next_col(cur_max: int) -> int:
        return cur_max + 1

    cur = last_header_col
    if not changed_col:
        cur = next_col(cur); changed_col = cur
    if not status_col:
        cur = next_col(cur); status_col = cur
    if not newname_col:
        cur = next_col(cur); newname_col = cur
    if not notes_col:
        cur = next_col(cur); notes_col = cur

    existing_audit_cols = any(k in existing_map for k in ("file name ok","status","new file name","notes"))

    return dict(
        header_row=header_row,
        file_name_col=file_col,
        company_col=company_col,
        out_header_row=header_row,
        changed_col=changed_col,
        status_col=status_col,
        newname_col=newname_col,
        notes_col=notes_col,
        existing_audit_columns=existing_audit_cols
    )

# ---------- Decision helper ----------
def assemble_decision(suggested_or_none: Optional[str], issues: List[str]) -> Tuple[Optional[str], List[str], bool]:
    if any("Does not follow CTA" in s for s in issues):
        return "", issues, True
    if suggested_or_none is None:
        return "", issues, True
    return suggested_or_none, issues, True

# ---------- Core audit ----------
def audit_filename(original_name: str) -> Tuple[Optional[str], List[str], bool]:
    original = original_name.strip()
    issues: List[str] = []
    if not original:
        return "", ["Empty value"], True

    for ig in IGNORE_PREFIXES:
        if original.startswith(ig):
            return None, [], False

    # Space before extension?
    if SPACE_BEFORE_EXT.search(original_name):
        issues.append("Removed space before file extension.")

    # stem/ext
    if "." in original:
        stem = original[:original.rfind(".")]
        ext  = original[original.rfind("."):]
    else:
        stem, ext = original, ""

    # compress double underscores first
    if "__" in stem:
        stem = MULTI_UNDERS.sub("_", stem)
        issues.append("Removed double underscore.")

    tokens = [t for t in stem.split("_") if t != ""]
    if len(tokens) < 2:
        issues.append("Does not follow CTA file naming convention.")
        return assemble_decision(None, issues)

    prefix = tokens[0]
    rest = tokens[1:]

    if not re.match(r"^[A-Z]{2,4}$", prefix):
        issues.append("Does not follow CTA file naming convention.")
        return assemble_decision(None, issues)

    def finalize(prefix: str, id_part: Optional[str], date_token: str, title_raw: str, hard_rev: Optional[str]) -> Tuple[str, List[str], bool]:
        title_no_rev, rev_norm = extract_rev(title_raw, issues)
        rev = hard_rev if hard_rev else rev_norm
        id_clean = id_part.replace("_", "") if id_part else None
        title = sanitize_title(title_no_rev, issues)
        title = enforce_stem_limit(prefix, id_clean, date_token, title, rev, issues)
        suggested_stem = join_tokens(prefix, id_clean, date_token, title, rev)
        if not underscores_ok(suggested_stem, bool(id_clean)):
            issues.append("Unexpected underscore count after normalization (check for missing title or extra separators).")
        suggested_stem = postprocess_before_ext(suggested_stem, issues)
        return suggested_stem + ext, issues, (suggested_stem + ext != original)

    def need_date(tok: Optional[str]) -> Tuple[Optional[str], bool]:
        if not tok: return None, False
        d, note = try_normalize_date(tok)
        if d and note: issues.append(note)
        return d, bool(d)

    # COR: date immediately after prefix
    if prefix == "COR":
        idx_date = None
        for i, tok in enumerate(rest):
            d, ok = need_date(tok)
            if ok:
                idx_date = i; date_token = d; break
        if idx_date is None:
            issues.append("Invalid or missing date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        if idx_date != 0:
            issues.append("Removed non-date token(s) before date.")
        title_raw = "_".join(rest[idx_date+1:])
        return finalize(prefix, None, date_token, title_raw, None)

    # QA: optional ID (any)
    if prefix == "QA":
        if not rest:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if ok0:
            title_raw = "_".join(rest[1:])
            return finalize(prefix, None, d0, title_raw, None)
        idp = sanitize_id_any(rest[0], issues, "QMS-ID")
        if len(rest) < 2:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d1, ok1 = need_date(rest[1])
        if not ok1:
            issues.append("Invalid or missing date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        title_raw = "_".join(rest[2:])
        return finalize(prefix, idp, d1, title_raw, None)

    # AU: requires any-chars ID
    if prefix == "AU":
        if len(rest) < 2:
            issues.append("Missing Audit-ID.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if ok0:
            issues.append("Missing Audit-ID before date. Expected AU_[Audit-ID]_[Date]_[Title]_Rev##.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        idp = sanitize_id_any(rest[0], issues, "Audit-ID")
        d1, ok1 = need_date(rest[1])
        if not ok1:
            issues.append("Invalid or missing date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        title_raw = "_".join(rest[2:])
        return finalize(prefix, idp, d1, title_raw, None)

    # FDM: TaskID[_ContractID] then date
    if prefix == "FDM":
        date_idx = None
        date_token = None
        for i, tok in enumerate(rest):
            d, ok = need_date(tok)
            if ok:
                date_idx = i; date_token = d; break
        if date_idx is None or date_idx == 0:
            issues.append("Missing TaskID/ContractID and/or date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        id_tokens = rest[:date_idx][:2]
        if not id_tokens:
            issues.append("Missing TaskID/ContractID.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        id_sanitized = [sanitize_id_any(t, issues, "ID") for t in id_tokens]
        id_part = "-".join([t for t in id_sanitized if t])
        title_raw = "_".join(rest[date_idx+1:])
        return finalize(prefix, id_part, date_token, title_raw, None)

    # ---------- TF override (approved) ----------
    # Allow missing FormID: TF_[Date]_[Title]_Rev##,
    # or reinterpret first token as Title when remainder is just Rev/copy index.
    if prefix == "TF":
        if not rest:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)

        # Case 1: Date immediately after TF → treat as missing FormID (allowed)
        d0, ok0 = need_date(rest[0])
        if ok0:
            title_raw = "_".join(rest[1:])
            issues.append("No FormID provided — allowed for TF.")
            return finalize(prefix, None, d0, title_raw, None)

        # Case 2: First token is Title, second is Date, remainder just Rev/copy → reinterpret
        d1, ok1 = need_date(rest[1] if len(rest) > 1 else None)
        if ok1:
            trailing = "_".join(rest[2:]) if len(rest) > 2 else ""
            # Try to capture digits from trailing rev; default Rev0
            m = REV_ANY_INFIX.search(trailing) or EMBEDDED_REV.search(trailing)
            digits = (m.group(1) if m else "0")
            digits = digits.upper().replace("O", "0")[:2] if digits else "0"
            # Drop trailing copy index like "(2)" if present
            trailing = re.sub(r"\s*\(\d+\)\s*$", "", trailing)
            issues.append("Reinterpreted first token as Title — missing FormID allowed for TF.")
            issues.append("Normalized revision element.")
            return finalize(prefix, None, d1, rest[0], f"Rev{digits}")
        # Otherwise fall through to generic REQUIRES_ID_ANY handling (TF with FormID)

    # Requires any-chars ID (guard against date-first)
    if prefix in REQUIRES_ID_ANY:
        if len(rest) < 2:
            issues.append(f"Missing {ID_LABEL.get(prefix,'ID')} and/or date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if ok0:
            lbl = ID_LABEL.get(prefix, "ID")
            issues.append(f"Missing {lbl} before date. Expected {prefix}_[{lbl}]_[Date]_[Title]_Rev##.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        idp = sanitize_id_any(rest[0], issues, ID_LABEL.get(prefix, "ID"))
        d1, ok1 = need_date(rest[1])
        if not ok1:
            issues.append("Invalid or missing date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        title_raw = "_".join(rest[2:])
        return finalize(prefix, idp, d1, title_raw, None)

    # ---------- CSI-required group (updated per request) ----------
    if prefix in REQUIRES_CSI:
        if len(rest) < 2:
            issues.append("Missing CSI section.")
            return assemble_decision(None, issues)

        # If the first token is actually a DATE, then CSI is missing.
        d_datefirst, ok_datefirst = need_date(rest[0])
        if ok_datefirst:
            issues.append("Missing CSI section.")
            return assemble_decision(None, issues)

        csi, cnote = normalize_csi(rest[0])
        if not csi:
            issues.append("Invalid CSI-Section (require NN-NN-NN or NN-NN-NN-NN).")
            return assemble_decision(None, issues)
        if cnote: issues.append(cnote)
        d1, ok1 = need_date(rest[1])
        if not ok1:
            issues.append("Invalid or missing date.")
            return assemble_decision(None, issues)
        title_raw = "_".join(rest[2:])
        return finalize(prefix, csi, d1, title_raw, None)

    # PN: optional CSI (if present must be valid)
    if prefix == "PN":
        if not rest:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if ok0:
            title_raw = "_".join(rest[1:])
            return finalize(prefix, None, d0, title_raw, None)
        csi, cnote = normalize_csi(rest[0])
        if csi:
            d1, ok1 = need_date(rest[1] if len(rest)>1 else None)
            if not ok1:
                issues.append("Invalid or missing date.")
                issues.append("Does not follow CTA file naming convention.")
                return assemble_decision(None, issues)
            if cnote: issues.append(cnote)
            title_raw = "_".join(rest[2:])
            return finalize(prefix, csi, d1, title_raw, None)
        issues.append("Does not follow CTA file naming convention.")
        return assemble_decision(None, issues)

    # SP: optional CSI (if present must be valid)
    if prefix == "SP":
        if not rest:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if ok0:
            title_raw = "_".join(rest[1:])
            return finalize(prefix, None, d0, title_raw, None)
        csi, cnote = normalize_csi(rest[0])
        if csi:
            d1, ok1 = need_date(rest[1] if len(rest)>1 else None)
            if not ok1:
                issues.append("Invalid or missing date.")
                issues.append("Does not follow CTA file naming convention.")
                return assemble_decision(None, issues)
            if cnote: issues.append(cnote)
            title_raw = "_".join(rest[2:])
            return finalize(prefix, csi, d1, title_raw, None)
        issues.append("Does not follow CTA file naming convention.")
        return assemble_decision(None, issues)

    # WR: CSI optional; if omitted, allowed and noted
    if prefix == "WR":
        if not rest:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if ok0:
            issues.append("No CSI provided — allowed for overall warranty.")
            title_raw = "_".join(rest[1:])
            return finalize(prefix, None, d0, title_raw, None)
        csi, cnote = normalize_csi(rest[0])
        if csi:
            d1, ok1 = need_date(rest[1] if len(rest)>1 else None)
            if not ok1:
                issues.append("Invalid or missing date.")
                issues.append("Does not follow CTA file naming convention.")
                return assemble_decision(None, issues)
            if cnote: issues.append(cnote)
            title_raw = "_".join(rest[2:])
            return finalize(prefix, csi, d1, title_raw, None)
        issues.append("Does not follow CTA file naming convention.")
        return assemble_decision(None, issues)

    # RE: simple standard (no ID)
    if prefix == "RE":
        if not rest:
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        d0, ok0 = need_date(rest[0])
        if not ok0:
            issues.append("Invalid or missing date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        title_raw = "_".join(rest[1:])
        return finalize(prefix, None, d0, title_raw, None)

    # Standard (no ID)
    if prefix in STANDARD_PREFIXES:
        d0, ok0 = need_date(rest[0] if rest else None)
        if not ok0:
            issues.append("Invalid or missing date.")
            issues.append("Does not follow CTA file naming convention.")
            return assemble_decision(None, issues)
        title_raw = "_".join(rest[1:])
        return finalize(prefix, None, d0, title_raw, None)

    # Unknown
    issues.append("Does not follow CTA file naming convention.")
    return assemble_decision(None, issues)

# ---------- Writer ----------
def write_results_inline_dynamic(input_file: str, sheet_name: str | None):
    src = Path(input_file)
    if not src.exists(): raise FileNotFoundError(src)

    # .xls conversion via Excel (Windows)
    cleanup_temp = None
    wb_path = src
    if src.suffix.lower() == ".xls":
        try:
            import win32com.client as win32
            import pythoncom
        except Exception as e:
            raise RuntimeError("win32com is required for .xls. Install: pip install pywin32") from e
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        try:
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(src))
            tmp = Path(tempfile.gettempdir()) / (src.stem + "_CONVERTED.xlsx")
            wb.SaveAs(str(tmp), FileFormat=51, ConflictResolution=2)
            wb.Close(SaveChanges=False)
            wb_path = tmp
            cleanup_temp = tmp
        finally:
            excel.Quit()
            try: pythoncom.CoUninitialize()
            except Exception: pass

    wb = load_workbook(wb_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    layout = detect_layout(ws)
    header_row = layout["header_row"]
    out_row    = layout["out_header_row"]
    file_col   = layout["file_name_col"]
    company_col = layout["company_col"]
    changed_col = layout["changed_col"]
    status_col  = layout["status_col"]
    newname_col = layout["newname_col"]
    notes_col   = layout["notes_col"]

    # Headers (strings)
    header_font = Font(bold=True, size=12, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def set_header(r, c, text):
        cell = ws.cell(row=r, column=c, value=str(text))
        cell.font = header_font; cell.alignment = header_alignment; cell.fill = header_fill; cell.border = thin

    set_header(out_row, changed_col, "File Name OK")
    set_header(out_row, status_col,  "Status")
    set_header(out_row, newname_col, "New file name")
    set_header(out_row, notes_col,   "Notes")
    ws.row_dimensions[out_row].height = 25

    last_row = find_last_row(ws, file_col, header_row)

    stats = {'noncompliant': 0, 'compliant': 0, 'empty_files': 0, 'skipped_files': 0}
    by_company: Dict[str, Dict[str,int]] = {}

    def bump(company: str, key: str):
        company = company or "Unknown"
        d = by_company.setdefault(company, {'noncompliant':0, 'compliant':0, 'empty_files':0, 'skipped_files':0})
        d[key] += 1

    center = Alignment(horizontal='center', vertical='center')
    wrap   = Alignment(wrap_text=True, vertical='top')

    for r in range(header_row + 1, last_row + 1):
        raw = ws.cell(row=r, column=file_col).value
        filename = safe_str(raw)
        company = safe_str(ws.cell(row=r, column=company_col).value) if company_col else "Unknown"

        dot_cell    = ws.cell(row=r, column=changed_col, value=DOT_CHAR)
        status_cell = ws.cell(row=r, column=status_col)

        if not filename:
            dot_cell.alignment = center
            dot_cell.font = RED_DOT_FONT
            status_cell.value = "Needs Changes"
            ws.cell(row=r, column=notes_col, value="Empty value")
            stats['empty_files'] += 1; bump(company, 'empty_files')
            for c in (changed_col, status_col, newname_col, notes_col):
                ws[f"{get_column_letter(c)}{r}"].border = thin
            ws[f"{get_column_letter(notes_col)}{r}"].alignment = wrap
            continue

        if any(filename.startswith(ig) for ig in IGNORE_PREFIXES):
            stats['skipped_files'] += 1; bump(company, 'skipped_files')
            continue

        suggestion, notes, _ = audit_filename(filename)
        if suggestion is None:
            stats['skipped_files'] += 1; bump(company, 'skipped_files')
            continue

        needs_attention = bool(notes) or (bool(suggestion) and suggestion != filename)

        dot_cell.alignment = center
        dot_cell.font = (RED_DOT_FONT if needs_attention else GREEN_DOT_FONT)
        status_cell.value = ("Needs Changes" if needs_attention else "OK")

        ws.cell(row=r, column=newname_col, value=(suggestion if (suggestion and suggestion != filename) else ""))
        ws.cell(row=r, column=notes_col,   value=("; ".join(notes) if notes else ""))

        ws[f"{get_column_letter(newname_col)}{r}"].alignment = wrap
        ws[f"{get_column_letter(notes_col)}{r}"].alignment = wrap
        for c in (changed_col, status_col, newname_col, notes_col):
            ws[f"{get_column_letter(c)}{r}"].border = thin

        if needs_attention:
            stats['noncompliant'] += 1; bump(company, 'noncompliant')
        else:
            stats['compliant'] += 1; bump(company, 'compliant')

    # Column widths
    for c in (changed_col, status_col, newname_col, notes_col):
        letter = get_column_letter(c)
        if c == changed_col:
            ws.column_dimensions[letter].width = 8
            continue
        if c == status_col:
            ws.column_dimensions[letter].width = 16
            continue
        maxlen = 0
        for rr in range(out_row, last_row + 1):
            v = ws[f"{letter}{rr}"].value
            if v is not None: maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = min(max(maxlen + 3, 12), 60)

    # Table (include Status col) — safe replace if exists
    table_ref = f"{get_column_letter(changed_col)}{out_row}:{get_column_letter(notes_col)}{last_row}"
    if any(t.displayName == "AuditResults" for t in ws._tables):
        ws._tables = [t for t in ws._tables if t.displayName != "AuditResults"]
    tbl = Table(displayName="AuditResults", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(tbl)

    analyzed_out = src.with_suffix(".xlsx").with_name(src.stem + "_ANALYZED.xlsx")
    wb.save(analyzed_out)

    if cleanup_temp and Path(cleanup_temp).exists():
        try: os.remove(cleanup_temp)
        except Exception: pass

    # Summary
    total_processed = stats['noncompliant'] + stats['compliant']
    total_files = total_processed + stats['empty_files'] + stats['skipped_files']
    print("\n" + "="*60)
    print("📊 CTA FILE NAMING AUDIT SUMMARY")
    print("="*60)
    print(f"📁 Total files analyzed: {total_files}")
    print(f"   • Files processed: {total_processed}")
    print(f"   • Empty/blank files: {stats['empty_files']}")
    print(f"   • Skipped files (I-PAY, PCO, IDR, NCR, RFI, SMTL, UOR): {stats['skipped_files']}")
    if total_processed > 0:
        print(f"   ❌ Noncompliant: {stats['noncompliant']} ({stats['noncompliant']/total_processed*100:.1f}% of processed)")
        print(f"   ✅ Compliant:    {stats['compliant']} ({stats['compliant']/total_processed*100:.1f}% of processed)")
    else:
        print("   ❌ Noncompliant: 0\n   ✅ Compliant:    0")
    print("\n🏢 Breakdown by Company")
    print("-"*60)
    for company, b in sorted(by_company.items(), key=lambda kv: (-kv[1]['noncompliant'], kv[0].lower())):
        proc = b['noncompliant'] + b['compliant']
        rate = (b['noncompliant']/proc*100) if proc else 0
        print(f"{company:30s}  ❌{b['noncompliant']:4d}  ✅{b['compliant']:4d}  ∅{b['empty_files']:3d}  ⏭ {b['skipped_files']:3d}   (noncomp {rate:4.1f}%)")
    print("="*60)

    # Include note if we reused existing audit columns
    if layout.get("existing_audit_columns"):
        print("ℹ️ Detected existing audit columns — updating rows in place.")

    return analyzed_out, layout

# ---------- Flask Web Service ----------
@app.route('/audit', methods=['POST'])
def audit_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Save uploaded file temporarily
    temp_dir = tempfile.mkdtemp()    
    input_path = os.path.join(temp_dir, file.filename)
    file.save(input_path)
    
    try:
        # Run your audit function
        output_path, layout = write_results_inline_dynamic(input_path, None)
        
        # Return the processed file
        return send_file(output_path, as_attachment=True, 
                        download_name=f"ANALYZED_{file.filename}")
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        # Cleanup
        try:
            if os.path.exists(input_path):
                os.remove(input_path)
        except:
            pass

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'ruleset': RULESET_NAME})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
